#!/usr/bin/env python3
"""Parse a RobStride Motor Assistant parameter export.

Motor Assistant writes `.xls` files that are actually `.xlsx` (OOXML/zip).
This script:
  1. Opens the workbook with openpyxl.
  2. Prints every sheet as a plaintext table (for audit/diffing).
  3. Emits a YAML-friendly summary of the parameters we care about for
     commissioning (safety limits, firmware version, can_status, etc.).

Usage:
    python tools/robstride/parse_motor_assistant_export.py <path-to-xls-or-xlsx>
"""

from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def _as_xlsx_path(path: Path) -> tuple[Path, Path | None]:
    """Motor Assistant saves OOXML (.xlsx) with an `.xls` extension; openpyxl
    refuses that. Return a usable path plus an optional tempfile to clean up
    after the caller closes the workbook.
    """
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return path, None
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copyfile(path, tmp.name)
        return Path(tmp.name), Path(tmp.name)

# Parameters we specifically care about during commissioning. Index hex is
# matched case-insensitively against the first column.
PARAMS_OF_INTEREST: dict[str, str] = {
    "0x1003": "AppCodeVersion",
    "0x1007": "AppCodeName",
    "0x2005": "MechOffset",
    "0x2007": "Status1 (factory torque limit)",
    "0x2009": "CAN_ID",
    "0x200A": "CAN_MASTER",
    "0x200B": "CAN_TIMEOUT",
    "0x2018": "limit_spd",
    "0x2019": "limit_cur",
    "0x2023": "zero_sta",
    "0x2024": "position_offset",
    "0x2026": "damper",
    "0x2027": "add_offset",
    "0x3041": "can_status",
    "0x3016": "mechPos",
    "0x3022": "faultSta",
    "0x700B": "limit_torque (7000-range mirror)",
    "0x7017": "limit_spd (7000-range mirror)",
    "0x7018": "limit_cur (7000-range mirror)",
    "0x7028": "canTimeout (7000-range)",
    "0x702A": "damper (7000-range)",
}


def _norm_index(s: object) -> str:
    return str(s).strip().lower() if s is not None else ""


def dump_workbook(path: Path) -> None:
    xlsx_path, tmp = _as_xlsx_path(path)
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"=== sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    if any(c.strip() for c in cells):
                        print("\t".join(cells))
                print()
        finally:
            wb.close()
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)


def summarize(path: Path) -> dict[str, tuple[str, str]]:
    """Return { '0x1003': (name, value_string) } for params of interest."""
    xlsx_path, tmp = _as_xlsx_path(path)
    found: dict[str, tuple[str, str]] = {}
    wanted = {k.lower(): v for k, v in PARAMS_OF_INTEREST.items()}
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    idx = _norm_index(row[0])
                    if idx in wanted:
                        value_cells = [c for c in row[1:] if c is not None]
                        value_str = " | ".join(str(c) for c in value_cells).strip()
                        found[idx] = (wanted[idx], value_str)
        finally:
            wb.close()
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)
    return found


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"error: {path} does not exist", file=sys.stderr)
        return 1

    print(f"# file: {path}")
    print(f"# size: {path.stat().st_size} bytes")
    print()

    dump_workbook(path)

    print("=== commissioning-relevant parameters ===")
    summary = summarize(path)
    if not summary:
        print("(none of the parameters-of-interest matched; check sheet format)")
    else:
        col_w = max(len(k) for k in summary)
        for idx, (name, val) in summary.items():
            print(f"{idx:<{col_w}}  {name:<40}  {val}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
